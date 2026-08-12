"""TableExtractor interface, registry, and per-format extractors (frozen §4.4).

Plugins implement :class:`TableExtractor` and register under the ``source_kinds``
they serve; the registry selects by kind and handles the empty-registry case.
Extractors are pure — they never raise; a failed extraction yields an empty list
so the enrichment stage falls back to flat text (frozen §2.4 failure modes).
"""

from __future__ import annotations

import csv
import io
import logging
from typing import Any, Protocol, cast, runtime_checkable

from app.domain.document_intelligence import Table, TableCell, TableHeader, TableRow
from app.domain.documents import SourceDocument

logger = logging.getLogger(__name__)

MAX_ROWS = 200  # frozen §2.4 default row cap (config-tunable)
MAX_COLS = 30  # frozen §2.4 default column cap (config-tunable)


@runtime_checkable
class TableExtractor(Protocol):
    """Extract structured tables from a source document (frozen §2.4)."""

    source_kinds: set[str]

    def extract(self, document: SourceDocument) -> list[Table]:
        """Return the tables found in ``document``; an empty list when none."""


class TableExtractorRegistry:
    """Select a table extractor by source kind (frozen §4.4 P2-402)."""

    def __init__(self, extractors: list[TableExtractor] | None = None) -> None:
        self._by_kind: dict[str, TableExtractor] = {}
        for extractor in extractors or []:
            self.register(extractor)

    def register(self, extractor: TableExtractor) -> None:
        for kind in extractor.source_kinds:
            self._by_kind[kind] = extractor

    def select(self, kind: str) -> TableExtractor | None:
        """Return the extractor for ``kind``, or ``None`` when unregistered."""
        return self._by_kind.get(kind)


def _cell(value: object) -> TableCell:
    """Normalize a raw cell value to text (empty string for None/blank)."""
    if value is None:
        return TableCell(value="")
    return TableCell(value=str(value))


def _build_table(
    *,
    title: str,
    raw_header: list[object],
    raw_rows: list[list[object]],
    source_position: str,
    max_rows: int,
    max_cols: int,
) -> Table | None:
    """Build a Table from raw header/data rows; None when nothing to render."""
    header = TableHeader(cells=[_cell(c) for c in raw_header[:max_cols]])
    rows = [
        TableRow(cells=[_cell(c) for c in row[:max_cols]])
        for row in raw_rows[:max_rows]
        if any(str(c).strip() for c in row)
    ]
    if not any(c.value.strip() for c in header.cells) and not rows:
        return None
    return Table(
        title=title,
        header=header,
        rows=rows,
        source_position=source_position,
    )


class CsvTableExtractor:
    """Extract tables from CSV/TSV text (frozen §4.4 P2-403)."""

    source_kinds = {"csv"}

    def __init__(
        self,
        *,
        header_sniffing: bool = True,
        max_rows: int = MAX_ROWS,
        max_cols: int = MAX_COLS,
    ) -> None:
        self._header_sniffing = header_sniffing
        self._max_rows = max_rows
        self._max_cols = max_cols

    def extract(self, document: SourceDocument) -> list[Table]:
        if not document.text:
            return []
        try:
            dialect = csv.Sniffer().sniff(document.text[:4096], delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        sample = io.StringIO(document.text)
        reader = csv.reader(sample, dialect)
        rows: list[list[str]] = []
        for row in reader:
            rows.append(row)
            if len(rows) >= self._max_rows + 1:
                break
        if not rows:
            return []
        raw_header: list[object] = list(rows[0]) if self._header_sniffing else []
        data_rows = rows[1:] if self._header_sniffing and rows else rows
        raw_rows: list[list[object]] = [list(r) for r in data_rows]
        table = _build_table(
            title=document.filename or "CSV Table",
            raw_header=raw_header,
            raw_rows=raw_rows,
            source_position="line 1",
            max_rows=self._max_rows,
            max_cols=self._max_cols,
        )
        return [table] if table is not None else []


class SpreadsheetTableExtractor:
    """Extract per-sheet tables from .xlsx workbooks (frozen §4.4 P2-404).

    The workbook is loaded **non-read-only** (``read_only=False``,
    ``data_only=True``) so ``merged_cells.ranges`` is available for flattening
    (R1). The ingestor's flat-text pass keeps its existing read-only pattern.
    """

    source_kinds = {"spreadsheet"}

    def __init__(
        self,
        *,
        header_sniffing: bool = True,
        max_rows: int = MAX_ROWS,
        max_cols: int = MAX_COLS,
    ) -> None:
        self._header_sniffing = header_sniffing
        self._max_rows = max_rows
        self._max_cols = max_cols

    def extract(self, document: SourceDocument) -> list[Table]:
        if document.source_path is None:
            return []
        try:
            import openpyxl  # type: ignore[import-untyped]

            wb = openpyxl.load_workbook(str(document.source_path), read_only=False, data_only=True)
        except ImportError:
            logger.warning("openpyxl is required for spreadsheet table extraction.")
            return []
        except Exception:
            logger.warning(
                "Spreadsheet table extraction failed.",
                extra={"source": document.source},
                exc_info=True,
            )
            return []
        try:
            tables: list[Table] = []
            for sheet_index, sheet_name in enumerate(wb.sheetnames):
                ws = wb[sheet_name]
                table = self._extract_sheet(ws, sheet_name, sheet_index + 1)
                if table is not None:
                    tables.append(table)
            return tables
        finally:
            wb.close()

    def _extract_sheet(self, ws: Any, sheet_name: str, sheet_index: int) -> Table | None:
        merged = (
            [(str(rng), _range_bounds(rng)) for rng in ws.merged_cells.ranges]
            if ws.merged_cells
            else []
        )
        matrix: list[list[object]] = []
        for row in ws.iter_rows(values_only=True):
            matrix.append(list(row))
            if len(matrix) >= self._max_rows + 1:
                break
        matrix = _flatten_merged_cells(matrix, merged)
        matrix = [row for row in matrix if any(c is not None and str(c).strip() for c in row)]
        if not matrix:
            return None
        raw_header = matrix[0] if self._header_sniffing else []
        raw_rows = matrix[1:] if self._header_sniffing else matrix
        return _build_table(
            title=sheet_name,
            raw_header=raw_header,
            raw_rows=raw_rows,
            source_position=f"sheet {sheet_index}",
            max_rows=self._max_rows,
            max_cols=self._max_cols,
        )


def _range_bounds(rng: object) -> tuple[int, int, int, int]:
    """Return (min_row, max_row, min_col, max_col) for an openpyxl range."""
    cell_refs = str(rng).split(":")
    starts = [_split_cell_ref(cell_ref) for cell_ref in cell_refs]
    min_row = min(row for row, _ in starts)
    max_row = max(row for row, _ in starts)
    min_col = min(col for _, col in starts)
    max_col = max(col for _, col in starts)
    return min_row, max_row, min_col, max_col


def _split_cell_ref(cell_ref: str) -> tuple[int, int]:
    """Split an A1-style reference into (row_index, col_index), both 0-based."""
    return _row_index(cell_ref), _col_index(cell_ref)


def _flatten_merged_cells(
    matrix: list[list[object]],
    merged_ranges: list[tuple[str, tuple[int, int, int, int]]],
) -> list[list[object]]:
    """Propagate the top-left value across each merged-cell rectangle (R1).

    Every cell inside a merged rectangle that is not the top-left anchor
    receives the anchor's value (only when it is currently None), so no data
    is silently lost when the sheet is normalized to a grid.
    """
    for _range_str, (min_row, max_row, min_col, max_col) in merged_ranges:
        if not (0 <= min_row < len(matrix)):
            continue
        anchor = matrix[min_row][min_col] if min_col < len(matrix[min_row]) else None
        for row in range(min_row, min(max_row, len(matrix) - 1) + 1):
            for col in range(min_col, max_col + 1):
                if row >= len(matrix) or col >= len(matrix[row]):
                    continue
                if row == min_row and col == min_col:
                    continue
                if matrix[row][col] is None:
                    matrix[row][col] = anchor
    return matrix


def _col_index(cell_ref: str) -> int:
    col_letters = "".join(ch for ch in cell_ref if ch.isalpha())
    index = 0
    for ch in col_letters:
        index = index * 26 + (ord(ch.upper()) - ord("A") + 1)
    return index - 1


def _row_index(cell_ref: str) -> int:
    digits = "".join(ch for ch in cell_ref if ch.isdigit())
    return (int(digits) - 1) if digits else -1


class PdfTableExtractor:
    """Extract tables from PDFs (frozen §4.4 P2-405).

    Default engine is pdfplumber (ADR-002). When ``pdf_engine="camelot"``,
    camelot is attempted and falls back to pdfplumber when unavailable or on
    failure. Engine missing → empty list (flat fallback + logged warning).
    pdfplumber exposes no per-table confidence in ``extract_tables`` output,
    so the frozen §2.4 ``min_confidence`` knob was removed (review R1 — dead
    config; see the M2.4 remediation report for the recorded deviation).
    """

    source_kinds = {"pdf"}

    def __init__(
        self,
        *,
        pdf_engine: str = "pdfplumber",
        max_rows: int = MAX_ROWS,
        max_cols: int = MAX_COLS,
    ) -> None:
        self._pdf_engine = pdf_engine
        self._max_rows = max_rows
        self._max_cols = max_cols

    def extract(self, document: SourceDocument) -> list[Table]:
        if document.source_path is None:
            return []
        if self._pdf_engine == "camelot":
            tables = self._extract_camelot(document)
            if tables:
                return tables
        return self._extract_pdfplumber(document)

    def _extract_pdfplumber(self, document: SourceDocument) -> list[Table]:
        try:
            import pdfplumber  # type: ignore[import-not-found]
        except ImportError:
            logger.warning("pdfplumber is required for PDF table extraction.")
            return []
        tables: list[Table] = []
        try:
            with pdfplumber.open(str(document.source_path)) as pdf:
                for page_index, page in enumerate(pdf.pages, start=1):
                    found = page.extract_tables()
                    for table_index, raw in enumerate(found, start=1):
                        # pdfplumber's stub widens cells to list[object]; they
                        # are str | None at runtime (a subset of object).
                        cells = cast(list[list[object]], raw)
                        table = _build_table(
                            title=f"Page {page_index} Table {table_index}",
                            raw_header=cells[0] if cells else [],
                            raw_rows=cells[1:] if cells else [],
                            source_position=f"page {page_index}",
                            max_rows=self._max_rows,
                            max_cols=self._max_cols,
                        )
                        if table is not None:
                            tables.append(table)
        except Exception:
            logger.warning(
                "PDF table extraction failed.",
                extra={"source": document.source},
                exc_info=True,
            )
            return []
        return tables

    def _extract_camelot(self, document: SourceDocument) -> list[Table]:
        try:
            import camelot  # type: ignore[import-not-found]
        except ImportError:
            logger.warning(
                "camelot not installed; falling back to pdfplumber for PDF tables."
            )
            return []
        tables: list[Table] = []
        try:
            parsed = camelot.read_pdf(str(document.source_path), pages="all")
            for table_index, frame in enumerate(parsed, start=1):
                values = frame.df.values.tolist()
                if not values:
                    continue
                table = _build_table(
                    title=f"Table {table_index}",
                    raw_header=values[0],
                    raw_rows=values[1:],
                    source_position=f"page {frame.page}",
                    max_rows=self._max_rows,
                    max_cols=self._max_cols,
                )
                if table is not None:
                    tables.append(table)
        except Exception:
            logger.warning(
                "camelot table extraction failed; falling back to pdfplumber.",
                extra={"source": document.source},
                exc_info=True,
            )
            return []
        return tables


def get_table_extractor(
    *,
    pdf_engine: str = "pdfplumber",
    max_rows: int = MAX_ROWS,
    max_cols: int = MAX_COLS,
    header_sniffing: bool = True,
) -> TableExtractorRegistry:
    """Composition root honoring TableSettings-style knobs (frozen §2.4)."""
    return TableExtractorRegistry(
        [
            CsvTableExtractor(
                header_sniffing=header_sniffing,
                max_rows=max_rows,
                max_cols=max_cols,
            ),
            SpreadsheetTableExtractor(
                header_sniffing=header_sniffing,
                max_rows=max_rows,
                max_cols=max_cols,
            ),
            PdfTableExtractor(
                pdf_engine=pdf_engine,
                max_rows=max_rows,
                max_cols=max_cols,
            ),
        ]
    )


def get_default_table_extractor() -> TableExtractorRegistry:
    """Return the default registry with the built-in extractors (composition root)."""
    return get_table_extractor()


def extract_tables(document: SourceDocument) -> list[Table]:
    """Extract tables from a document using the default registry (frozen public API)."""
    kind = document.source_type
    extractor = get_default_table_extractor().select(kind)
    if extractor is None:
        return []
    return extractor.extract(document)
