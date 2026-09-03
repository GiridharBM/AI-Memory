"""Spreadsheet (XLS/XLSX) file ingestor."""

from __future__ import annotations

from pathlib import Path

from app.domain.documents import DocumentMetadata, SourceDocument
from app.infrastructure.ingestion.base import (
    BaseIngestor,
    IngestionError,
    SourceReference,
    require_path_source,
)
from app.infrastructure.ingestion.utils import file_timestamp


class SpreadsheetIngestor(BaseIngestor):
    """Read spreadsheet files into normalized source documents."""

    source_type = "spreadsheet"
    supported_suffixes = (".xlsx", ".xls", ".ods")

    def ingest(self, source: SourceReference) -> SourceDocument:
        source_path = require_path_source(source, ingestor_name="Spreadsheet ingestor")
        text = self._extract_text(source_path)
        resolved_path = source_path.resolve()
        suffix = source_path.suffix.lower()
        if suffix == ".ods":
            mime = "application/vnd.oasis.opendocument.spreadsheet"
        elif suffix == ".xls":
            mime = "application/vnd.ms-excel"
        else:
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return SourceDocument(
            source=str(resolved_path),
            source_path=resolved_path,
            source_type=self.source_type,
            filename=source_path.name,
            text=text,
            metadata=DocumentMetadata(
                title=source_path.stem,
                modified_at=file_timestamp(source_path),
                mime_type=mime,
            ),
        )

    def _extract_text(self, path: Path) -> str:
        suffix = path.suffix.lower()
        if suffix == ".ods":
            return self._extract_ods(path)
        if suffix == ".xls":
            return self._extract_xls(path)
        return self._extract_xlsx(path)

    def _extract_xlsx(self, path: Path) -> str:
        try:
            import openpyxl  # type: ignore[import-untyped]

            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            texts: list[str] = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                texts.append(f"Sheet: {sheet_name}")
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(
                        str(cell) if cell is not None else "" for cell in row
                    )
                    if row_text.strip():
                        texts.append(row_text)
            wb.close()
            return "\n".join(texts)
        except ImportError:
            raise IngestionError(
                "openpyxl is required for spreadsheet ingestion. "
                "Install with: pip install openpyxl"
            ) from None
        except Exception as exc:
            raise IngestionError(f"Unable to read spreadsheet file '{path}'.") from exc

    def _extract_ods(self, path: Path) -> str:
        try:
            from odf import table, teletype  # type: ignore[import-untyped]
            from odf import text as odf_text
            from odf.opendocument import load  # type: ignore[import-untyped]

            doc = load(str(path))
            texts: list[str] = []
            for tbl in doc.getElementsByType(table.Table):
                sheet_name = tbl.getAttribute("name") or "Sheet"
                texts.append(f"Sheet: {sheet_name}")
                for row in tbl.getElementsByType(table.TableRow):
                    cells: list[str] = []
                    for cell in row.getElementsByType(table.TableCell):
                        repeat = cell.getAttribute("numbercolumnsrepeated")
                        if repeat and int(repeat) > 10:
                            continue
                        parts = [
                            teletype.extractText(p)
                            for p in cell.getElementsByType(odf_text.P)
                        ]
                        cells.append(" ".join(parts) if parts else "")
                    row_text = " | ".join(cells)
                    if row_text.strip():
                        texts.append(row_text)
            return "\n".join(texts)
        except ImportError:
            raise IngestionError(
                "odfpy is required for ODS ingestion. Install with: pip install odfpy"
            ) from None
        except IngestionError:
            raise
        except Exception as exc:
            raise IngestionError(f"Unable to read ODS file '{path}'.") from exc

    def _extract_xls(self, path: Path) -> str:
        try:
            import xlrd  # type: ignore[import-untyped]

            wb = xlrd.open_workbook(str(path))
            texts: list[str] = []
            for sheet in wb.sheets():
                texts.append(f"Sheet: {sheet.name}")
                for row_idx in range(sheet.nrows):
                    cells = [
                        str(sheet.cell_value(row_idx, col))
                        for col in range(sheet.ncols)
                    ]
                    row_text = " | ".join(cells)
                    if row_text.strip():
                        texts.append(row_text)
            return "\n".join(texts)
        except ImportError:
            raise IngestionError(
                "xlrd is required for XLS ingestion. Install with: pip install xlrd"
            ) from None
        except IngestionError:
            raise
        except Exception as exc:
            raise IngestionError(f"Unable to read XLS file '{path}'.") from exc
